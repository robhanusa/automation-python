# -*- coding: utf-8 -*-

from datetime import date, datetime, timedelta
import re
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting import Rule

upper_threshold = 20
middle_threshold = 10
today = datetime.today()
days_in_last_month = (date.today().replace(day=1)-timedelta(days=1)).day
month_timedelta = timedelta(days=days_in_last_month)
year_timedelta = timedelta(days = 366)

#create dictionaries with counts of FM's for the last month and year
wb = openpyxl.load_workbook('inputxl.xlsx', data_only=True)
ws = wb.active
def failurePercent(timeperiod:timedelta()) -> dict:
    FM_total = 0
    FM_count = {}
    FM_percent = {}
    for x in range(1,ws.max_row+1):
        FM = ws.cell(row=x, column=1).internal_value
        date = ws.cell(row=x, column=2).internal_value 
        if today - date <= timeperiod:
            FM_total += 1
            if FM not in FM_count.keys():
                FM_count[FM] = 1
            else:
                FM_count[FM] += 1
    #turn the failure counts into percentages
    for failure in FM_count.keys():
        FM_percent[failure] = round(FM_count[failure]/FM_total*100)
    return FM_percent

#look up failure mode severities via xlsx table
severities = {}
wb_severities = openpyxl.load_workbook('Severities and Occurrences.xlsx')
ws_severities = wb_severities['severities']
for x in range(1,ws_severities.max_row+1):
    severities[ws_severities.cell(row=x,column=1).internal_value] = ws_severities.cell(row=x,column=2).internal_value

#convert percent occurrences to occurrence scores via xlsx table of occurrence scores
occurrence_table = {}
ws_occurrences = wb_severities['occurrences']
for x in range(1,ws_occurrences.max_row+1):
    occurrence_table[ws_occurrences.cell(row=x,column=1).internal_value] = ws_occurrences.cell(row=x,column=2).internal_value
    
def occurrenceScore(timeperiod:timedelta(), occurrence_table: dict) -> dict:
    pattern_occurrence = r'(\d+) - (\d+)\%'
    FM_percent = failurePercent(timeperiod)
    occurrence_scores = {}
    for failure in FM_percent.items():
            for criteria in occurrence_table.items():
                crit = re.search(pattern_occurrence, criteria[0])
                if int(crit.groups()[0]) <= FM_percent[failure[0]] <= int(crit.groups()[1]):
                    occurrence_scores[failure[0]] = occurrence_table[criteria[0]]
    return occurrence_scores
                    
#Make dict of dict's that has RPNand other values we want in the final report
#here is where I'll call the above functions with 1 month and 1 year as input
rpn = {}
occurrences_month = occurrenceScore(month_timedelta, occurrence_table)
occurrences_year =  occurrenceScore(year_timedelta, occurrence_table) 
failure_percent_month = failurePercent(month_timedelta)
failure_percent_year = failurePercent(year_timedelta)

for failure in occurrences_year.keys():
    if failure not in occurrences_month:
        occurrences_month[failure] = 0
        failure_percent_month[failure] = 0
    rpn[failure]={'Monthly Failure Rate':f'{failure_percent_month[failure]}%',
                       'Yearly Failure Rate':f'{failure_percent_year[failure]}%',
                       'Monthly Occurrence': occurrences_month[failure],
                       'Yearly Occurrence': occurrences_year[failure],
                       'Severity': severities[failure],
                       'Monthly RPN' : occurrences_month[failure]*severities[failure],
                       'Yearly RPN': occurrences_year[failure]*severities[failure]}
    
    
#output xlsx file
report = openpyxl.Workbook()
ws = report.active

# create headers in first row
ws['A1'] = 'Failure Mode'
ws['B1'] = 'Yearly RPN'
ws['C1'] = 'Monthly RPN'
ws['D1'] = 'Severity'
ws['E1'] = 'Yearly Occurrence'
ws['F1'] = 'Monthly Occurrence'
ws['G1'] = 'Yearly Failure Rate'
ws['H1'] = 'Monthly Failure Rate'

header_list = []
for col in range(1, ws.max_column+1):
    header_list.append(ws.cell(1,col).value)

#populate cells with rpn dictionary contents
for x, failure in enumerate(rpn.keys()):
    ws.cell(row = x+2, column = 1, value = failure)
    for y in range(2,len(header_list)+1): 
        ws.cell(row = x+2, column = y, value = rpn[failure][header_list[y-1]])

#format column widths
ws.column_dimensions['A'].width = 25
for y in range(2,ws.max_column+1):
    i = get_column_letter(y)
    ws.column_dimensions[i].width = 12

#format all cells
for y in range(1,ws.max_column+1):
    ws.cell(row=1,column=y).font = Font(bold = True, size = 11)
    ws.cell(row=1,column=y).alignment = Alignment(wrap_text=True,vertical='top') 
    for x in range(2,ws.max_row+1):
        if y !=1:
            ws.cell(row=x,column=y).alignment = Alignment(horizontal='right')
 
#apply conditional formatting to rpn cells
last_cell = f'C{ws.max_row}'
redFill = PatternFill(bgColor = 'ffcccb')
redText = Font(color = '650000')
dxfredFill = DifferentialStyle(fill=redFill, font = redText)
greenText = Font(color="006100")
greenFill = PatternFill(bgColor="C6EFCE")
dxfgreen = DifferentialStyle(font = greenText, fill = greenFill)
yellowFill = PatternFill(bgColor='FFFF99')
yellowText = Font(color = '666600')
dxfyellowFill = DifferentialStyle(fill=yellowFill, font = yellowText)

rRed = Rule(type="expression", dxf=dxfredFill, stopIfTrue=True)
rRed.formula = [f'B2>={upper_threshold}']
ws.conditional_formatting.add(f"B2:{last_cell}", rRed)

rYellow = Rule(type="expression", dxf=dxfyellowFill, stopIfTrue=True)
rYellow.formula = [f'B2>={middle_threshold}']
ws.conditional_formatting.add(f"B2:{last_cell}", rYellow)

rGreen = Rule(type="expression", dxf=dxfgreen, stopIfTrue=True)
rGreen.formula = ['B2>=0']
ws.conditional_formatting.add(f"B2:{last_cell}", rGreen)

#create output report
report.save(filename = 'report.xlsx')
